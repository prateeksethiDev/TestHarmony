import openpyxl


class CheckoutPageData:
    test_checkout_page_data = [{"firstname": "John", "lastname": "Smith", "zipcode": 1234},
                               {"firstname": "Shawn", "lastname": "Smith", "zipcode": 3456}]

    @staticmethod
    def get_test_data(test_case_name):
        book = openpyxl.load_workbook("Provide Excel File Path")
        sheet = book.active
        Dict = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1) == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
